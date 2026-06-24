"""Data loaders for the Neal's CNC converter."""

import csv
import json
from pathlib import Path
from typing import Dict, List


def load_seed_records(seed_path, models=('core.user', 'core.accountingcategory',
                                         'jobs.serviceprice')):
    """Return records for `models` from a Django fixture JSON, verbatim.

    Each record is returned as its original dict (pk present or absent —
    nealseed's user records are written without explicit pks).
    """
    with open(seed_path, encoding='utf-8') as f:
        data = json.load(f)
    wanted = set(models)
    return [rec for rec in data if rec.get('model') in wanted]

try:
    import openpyxl
except ImportError:
    raise ImportError("openpyxl is required. Install with: pip install openpyxl")


def _single_dataset_file(matches, label, datasets_dir):
    """Return the sole file in `matches`, or raise ValueError if 0 or >1."""
    if not matches:
        raise ValueError(f'No {label} file found in {datasets_dir}')
    if len(matches) > 1:
        names = ', '.join(sorted(m.name for m in matches))
        raise ValueError(
            f'Expected exactly one {label} file in {datasets_dir}, '
            f'found {len(matches)}: {names}'
        )
    return matches[0]


def discover_datasets(datasets_dir):
    """Locate the single Excel file and single CSV file in `datasets_dir`.

    Returns (excel_path, csv_path) as strings. Raises ValueError with a
    clear message if either is missing or ambiguous (more than one).
    """
    datasets_dir = Path(datasets_dir)
    excels = sorted(datasets_dir.glob('*.xlsx')) + sorted(datasets_dir.glob('*.xls'))
    csvs = sorted(datasets_dir.glob('*.csv'))
    excel = _single_dataset_file(excels, 'Excel (.xlsx/.xls)', datasets_dir)
    csv_file = _single_dataset_file(csvs, 'CSV (.csv)', datasets_dir)
    return str(excel), str(csv_file)


class ExcelDataLoader:
    """Loads and parses Excel sheets into structured data."""

    def __init__(self, excel_path: str, verbose: bool = False):
        self.excel_path = excel_path
        self.verbose = verbose
        self.wb = None
        self.sheets_data = {}

    def load(self):
        """Load all required sheets into memory."""
        if self.verbose:
            print(f"Loading Excel file: {self.excel_path}")

        self.wb = openpyxl.load_workbook(self.excel_path, data_only=True)

        sheets_to_load = [
            'Contacts', 'Projects', 'Invoices', 'Estimates',
            'Bills', 'Price List Items'
        ]

        for sheet_name in sheets_to_load:
            if sheet_name in self.wb.sheetnames:
                self.sheets_data[sheet_name] = self._load_sheet(sheet_name)
                if self.verbose:
                    print(f"  Loaded {sheet_name}: {len(self.sheets_data[sheet_name])} rows")
            else:
                print(f"Warning: Sheet '{sheet_name}' not found in workbook")
                self.sheets_data[sheet_name] = []

        self.wb.close()

    def _load_sheet(self, sheet_name: str) -> List[Dict]:
        """Load sheet into list of dictionaries."""
        ws = self.wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))

        if not rows:
            return []

        headers = rows[0]
        data = []

        for row_idx, row_values in enumerate(rows[1:], start=2):
            row_dict = {
                '_row': row_idx,
                '_sheet': sheet_name
            }
            for idx, header in enumerate(headers):
                if header and idx < len(row_values):
                    row_dict[header] = row_values[idx]
            data.append(row_dict)

        return data


class KanbanCsvLoader:
    """Loads the Kanban board export (tab- or comma-delimited) into a list of dicts.

    The export format has flipped between tab- and comma-delimited across
    re-exports, so the delimiter is sniffed from the header line.
    """

    def __init__(self, csv_path):
        self.csv_path = csv_path

    def load(self):
        with open(self.csv_path, newline='', encoding='utf-8-sig') as f:
            first = f.readline()
            if first.lower().startswith('sep='):
                # explicit sep= directive — honour it and consume the line
                delimiter = first.strip()[len('sep='):] or '\t'
                header_line = f.readline()
            else:
                # No sep= directive — pick whichever of tab/comma the header
                # line has more of (header has 14+ fields; the loser will
                # typically have 0).
                delimiter = '\t' if first.count('\t') > first.count(',') else ','
                header_line = first
            reader = csv.DictReader(
                [header_line] + f.readlines(), delimiter=delimiter)
            return [dict(row) for row in reader]

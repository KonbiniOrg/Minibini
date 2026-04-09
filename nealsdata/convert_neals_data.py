#!/usr/bin/env python3
"""
Convert Neal's CNC Excel export to Django fixture JSON.

This script parses the Excel file, filters data to keep 100 Projects and related records,
handles contact name mismatches interactively, and generates a JSON fixture file for
import via Django's loaddata command.

Usage:
    python convert_neals_data.py nealsdata/company-export.xlsx
    python convert_neals_data.py nealsdata/company-export.xlsx --output my_data.json
    python convert_neals_data.py nealsdata/company-export.xlsx --non-interactive
    python convert_neals_data.py nealsdata/company-export.xlsx --dry-run
"""

import argparse
import json
import re
import sys
from datetime import date, datetime, timedelta
from decimal import Decimal
from pathlib import Path
from typing import Dict, List, Optional, Tuple, Any

try:
    import openpyxl
except ImportError:
    print("Error: openpyxl is required. Install with: pip install openpyxl")
    sys.exit(1)


class ContactMismatchHandler:
    """Handles interactive contact name mismatch resolution."""

    def __init__(self, interactive: bool = True):
        self.interactive = interactive
        self.decisions = {}  # Cache decisions for same mismatch

    def prompt_for_decision(
        self,
        business: str,
        existing_name: str,
        existing_email: str,
        referenced_name: str,
        sheet: str,
        row: int,
        context: str
    ) -> str:
        """
        Prompt user for decision on contact mismatch.
        Returns: 'update', 'create', or 'map'
        """
        key = (business, existing_name, referenced_name)

        # Return cached decision if we've seen this before
        if key in self.decisions:
            return self.decisions[key]

        # Non-interactive mode: always map to existing
        if not self.interactive:
            self.decisions[key] = 'map'
            return 'map'

        # Interactive prompt
        print("\n" + "=" * 64)
        print("Contact Mismatch Detected")
        print("=" * 64)
        print(f"Business:        {business}")
        print(f"Contacts Sheet:  {existing_name} ({existing_email})")
        print()
        print(f'Referenced as:   "{referenced_name}"')
        print(f"Found in:        {sheet} sheet, Row {row}")
        print(f"Context:         {context}")
        print()
        print("How should this be handled?")
        print(f'  [1] Update - Change contact name to "{referenced_name}"')
        print(f'  [2] Create - Create new contact "{referenced_name}" for this business')
        print(f'  [3] Map - Use existing contact "{existing_name}" as-is')
        print()

        while True:
            choice = input("Your choice (1/2/3): ").strip()
            if choice == '1':
                decision = 'update'
                break
            elif choice == '2':
                decision = 'create'
                break
            elif choice == '3':
                decision = 'map'
                break
            else:
                print("Invalid choice. Please enter 1, 2, or 3.")

        print("=" * 64)

        self.decisions[key] = decision
        return decision


class ExcelDataLoader:
    """Loads and parses Excel sheets into structured data."""

    def __init__(self, excel_path: str, verbose: bool = False):
        self.excel_path = excel_path
        self.verbose = verbose
        self.wb = None
        self.sheets_data = {}

    def load(self):
        """Load all required sheets into memory."""
        if self.verbose:
            print(f"Loading Excel file: {self.excel_path}")

        self.wb = openpyxl.load_workbook(self.excel_path, data_only=True)

        sheets_to_load = [
            'Contacts', 'Projects', 'Invoices', 'Estimates',
            'Bills', 'Tasks', 'Timeslips', 'Price List Items'
        ]

        for sheet_name in sheets_to_load:
            if sheet_name in self.wb.sheetnames:
                self.sheets_data[sheet_name] = self._load_sheet(sheet_name)
                if self.verbose:
                    print(f"  Loaded {sheet_name}: {len(self.sheets_data[sheet_name])} rows")
            else:
                print(f"Warning: Sheet '{sheet_name}' not found in workbook")
                self.sheets_data[sheet_name] = []

        self.wb.close()

    def _load_sheet(self, sheet_name: str) -> List[Dict]:
        """Load sheet into list of dictionaries."""
        ws = self.wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))

        if not rows:
            return []

        headers = rows[0]
        data = []

        for row_idx, row_values in enumerate(rows[1:], start=2):
            row_dict = {
                '_row': row_idx,
                '_sheet': sheet_name
            }
            for idx, header in enumerate(headers):
                if header and idx < len(row_values):
                    row_dict[header] = row_values[idx]
            data.append(row_dict)

        return data


class NealsDataConverter:
    """Main converter class that orchestrates the conversion process."""

    def __init__(
        self,
        excel_path: str,
        output_path: str = "neals_data.json",
        interactive: bool = True,
        dry_run: bool = False,
        verbose: bool = False
    ):
        self.excel_path = excel_path
        self.output_path = output_path
        self.interactive = interactive
        self.dry_run = dry_run
        self.verbose = verbose

        self.loader = ExcelDataLoader(excel_path, verbose)
        self.contact_handler = ContactMismatchHandler(interactive)

        # Data structures
        self.fixture_data = []
        self.pk_counters = {}  # Track next PK for each model

        # Lookup mappings
        self.business_map = {}  # org_name -> business_pk
        self.contact_map = {}  # (org_name, contact_name) -> contact_pk
        self.job_map = {}  # project_name -> (job_pk, workorder_pk or None)
        self.task_map = {}  # (project_name, task_name) -> task_pk

        # Filtered data
        self.filtered_contacts = []
        self.filtered_projects = []
        self.filtered_invoices = []
        self.filtered_estimates = []
        self.filtered_bills = []
        self.filtered_tasks = []
        self.filtered_timeslips = []
        self.filtered_price_list = []

        # Contact updates/creations
        self.contact_updates = {}  # pk -> new_name
        self.new_contacts = []  # List of new contact dicts to create

    def get_next_pk(self, model: str) -> int:
        """Get next primary key for a model type."""
        if model not in self.pk_counters:
            self.pk_counters[model] = 1
        pk = self.pk_counters[model]
        self.pk_counters[model] += 1
        return pk

    def add_fixture(self, model: str, pk: int, fields: Dict):
        """Add an object to the fixture data."""
        self.fixture_data.append({
            "model": model,
            "pk": pk,
            "fields": fields
        })

    def convert(self):
        """Main conversion process."""
        print("=" * 70)
        print("Neal's CNC Data Converter")
        print("=" * 70)
        print()

        # Phase 1: Load data
        print("[Phase 1] Loading Excel data...")
        self.loader.load()
        print()

        # Phase 2: Filter data
        print("[Phase 2] Filtering data...")
        self._filter_data()
        print()

        # Phase 3: Build objects (includes interactive contact resolution)
        print("[Phase 3] Building Django objects...")
        self._build_all_objects()
        print()

        # Summary
        self._print_summary()

        # Phase 4: Write JSON
        if not self.dry_run:
            self._write_json()
            print()
            print("=" * 70)
            print(f"✓ JSON fixture written to: {self.output_path}")
            print()
            print("To import into Django, run:")
            print(f"  python manage.py loaddata {self.output_path}")
            print("=" * 70)
        else:
            print()
            print("=" * 70)
            print("DRY RUN - No file written")
            print("=" * 70)

    def _filter_data(self):
        """Filter data to keep only relevant records."""
        # Keep all 100 Projects
        self.filtered_projects = self.loader.sheets_data.get('Projects', [])

        # Build set of project names for filtering
        project_names = {p.get('Name') for p in self.filtered_projects if p.get('Name')}

        if self.verbose:
            print(f"  Keeping {len(self.filtered_projects)} projects")
            print(f"  Project names set: {len(project_names)} unique names")

        # Filter other data based on project references
        # or has a recent date

        cutoff_date = datetime(2026, 1, 1)

        # Bills, Invoices and Estimates: Filter these first (they have line items mixed in)
        self.filtered_bills = self._filter_bills(project_names, cutoff_date)
        self.filtered_invoices = self._filter_invoices(project_names, cutoff_date)
        self.filtered_estimates = self._filter_estimates(project_names, cutoff_date)

        # Collect all organisations referenced by filtered data
        referenced_orgs = set()

        # From Projects
        for project in self.filtered_projects:
            org = project.get('Client Organisation')
            if org:
                referenced_orgs.add(org)

        # From Bills
        for bill in self.filtered_bills:
            org = bill.get('Contact Organisation')
            if org:
                referenced_orgs.add(org)

        # From Invoices
        for invoice in self.filtered_invoices:
            org = invoice.get('Contact Organisation')
            if org:
                referenced_orgs.add(org)

        # Also collect individual names referenced by projects (no org)
        referenced_names = set()
        for project in self.filtered_projects:
            if not project.get('Client Organisation'):
                name = project.get('Client Name')
                if name:
                    referenced_names.add(name)

        # Contacts: Keep those from referenced orgs OR referenced by name (no org)
        self.filtered_contacts = [
            c for c in self.loader.sheets_data.get('Contacts', [])
            if c.get('Organisation') in referenced_orgs
            or (not c.get('Organisation')
                and f"{c.get('First Name', '')} {c.get('Last Name', '')}".strip() in referenced_names)
        ]

        # Tasks: Keep those linked to kept projects
        self.filtered_tasks = [
            t for t in self.loader.sheets_data.get('Tasks', [])
            if t.get('Project') in project_names
        ]

        # Timeslips: Keep those linked to kept projects
        self.filtered_timeslips = [
            t for t in self.loader.sheets_data.get('Timeslips', [])
            if t.get('Project') in project_names
        ]

        # Price List Items: Keep all
        self.filtered_price_list = self.loader.sheets_data.get('Price List Items', [])

        if self.verbose:
            print(f"  Referenced organisations: {len(referenced_orgs)}")
            print(f"  Filtered contacts: {len(self.filtered_contacts)}")
            print(f"  Filtered tasks: {len(self.filtered_tasks)}")
            print(f"  Filtered timeslips: {len(self.filtered_timeslips)}")
            print(f"  Filtered bills: {len(self.filtered_bills)}")
            print(f"  Filtered invoices: {len(self.filtered_invoices)}")
            print(f"  Filtered estimates: {len(self.filtered_estimates)}")

    def _filter_invoices(self, project_names: set, cutoff_date: datetime) -> List[Dict]:
        """Filter invoice records (header + line items)."""
        invoices = []
        current_invoice = None

        for row in self.loader.sheets_data.get('Invoices', []):
            # Check if this is a header row (has Contact Organisation)
            if row.get('Contact Organisation'):
                # Save previous invoice if exists
                if current_invoice:
                    # Check if we should keep this invoice
                    projects = current_invoice.get('Projects', '')
                    if projects and any(p.strip() in project_names for p in projects.split(',')):
                        invoices.append(current_invoice)
                    elif current_invoice.get('Date') and isinstance(current_invoice.get('Date'), datetime) and current_invoice.get('Date') >= cutoff_date:
                        invoices.append(current_invoice)

                # Start new invoice
                current_invoice = row.copy()
                current_invoice['_line_items'] = []

            # Check if this is a line item row (has Item Type but no Contact Organisation)
            elif row.get('Item Type') and current_invoice:
                current_invoice['_line_items'].append(row)

        # Don't forget the last invoice
        if current_invoice:
            projects = current_invoice.get('Projects', '')
            if projects and any(p.strip() in project_names for p in projects.split(',')):
                invoices.append(current_invoice)
            elif current_invoice.get('Date') and isinstance(current_invoice.get('Date'), datetime) and current_invoice.get('Date') >= cutoff_date:
                invoices.append(current_invoice)

        return invoices

    def _filter_estimates(self, project_names: set, cutoff_date: datetime) -> List[Dict]:
        """Filter estimate records (header + line items)."""
        estimates = []
        current_estimate = None

        for row in self.loader.sheets_data.get('Estimates', []):
            # Check if this is a header row (has Reference - the estimate number)
            if row.get('Reference'):
                # Save previous estimate if exists
                if current_estimate:
                    project = current_estimate.get('Project', '')
                    if project and project in project_names:
                        estimates.append(current_estimate)
                    elif current_estimate.get('Date') and isinstance(current_estimate.get('Date'), datetime) and current_estimate.get('Date') >= cutoff_date:
                        estimates.append(current_estimate)

                # Start new estimate
                current_estimate = row.copy()
                current_estimate['_line_items'] = []

            # Check if this is a line item row (has Item Type but no Reference)
            elif row.get('Item Type') and current_estimate:
                current_estimate['_line_items'].append(row)

        # Don't forget the last estimate
        if current_estimate:
            project = current_estimate.get('Project', '')
            if project and project in project_names:
                estimates.append(current_estimate)
            elif current_estimate.get('Date') and isinstance(current_estimate.get('Date'), datetime) and current_estimate.get('Date') >= cutoff_date:
                estimates.append(current_estimate)

        return estimates

    def _is_bill_header_row(self, row: Dict) -> bool:
        """Detect bill header rows. A header has Contact Organisation, Contact Name, or Reference
        but NOT Item Type (which indicates a line item row)."""
        if row.get('Item Type'):
            return False
        return bool(row.get('Contact Organisation') or row.get('Contact Name') or row.get('Reference'))

    def _filter_bills(self, project_names: set, cutoff_date: datetime) -> List[Dict]:
        """Filter bill records (header + line items)."""
        bills = []
        current_bill = None

        for row in self.loader.sheets_data.get('Bills', []):
            if self._is_bill_header_row(row):
                # Save previous bill if exists
                if current_bill:
                    project = current_bill.get('Project', '')
                    if project in project_names:
                        bills.append(current_bill)
                    elif current_bill.get('Date') and isinstance(current_bill.get('Date'), datetime) and current_bill.get('Date') >= cutoff_date:
                        bills.append(current_bill)

                # Start new bill
                current_bill = row.copy()
                current_bill['_line_items'] = []

            # Line item row: has Item Type
            elif row.get('Item Type') and current_bill:
                current_bill['_line_items'].append(row)

        # Don't forget the last bill
        if current_bill:
            project = current_bill.get('Project', '')
            if project in project_names:
                bills.append(current_bill)
            elif current_bill.get('Date') and isinstance(current_bill.get('Date'), datetime) and current_bill.get('Date') >= cutoff_date:
                bills.append(current_bill)

        return bills

    def _collect_all_estimates(self) -> List[Dict]:
        """Parse ALL estimates from raw sheet data (with line items attached)."""
        estimates = []
        current = None

        for row in self.loader.sheets_data.get('Estimates', []):
            if row.get('Reference'):
                if current:
                    estimates.append(current)
                current = row.copy()
                current['_line_items'] = []
            elif row.get('Item Type') and current:
                current['_line_items'].append(row)

        if current:
            estimates.append(current)
        return estimates

    def _collect_all_invoices(self) -> List[Dict]:
        """Parse ALL invoices from raw sheet data (with line items attached)."""
        invoices = []
        current = None

        for row in self.loader.sheets_data.get('Invoices', []):
            if row.get('Contact Organisation') or (row.get('Contact Name') and row.get('Reference')):
                if current:
                    invoices.append(current)
                current = row.copy()
                current['_line_items'] = []
            elif row.get('Item Type') and current:
                current['_line_items'].append(row)

        if current:
            invoices.append(current)
        return invoices

    def _collect_all_bills(self) -> List[Dict]:
        """Parse ALL bills from raw sheet data (with line items attached)."""
        bills = []
        current = None

        for row in self.loader.sheets_data.get('Bills', []):
            if self._is_bill_header_row(row):
                if current:
                    bills.append(current)
                current = row.copy()
                current['_line_items'] = []
            elif row.get('Item Type') and current:
                current['_line_items'].append(row)

        if current:
            bills.append(current)
        return bills

    def _build_all_objects(self):
        """Build all Django objects in dependency order."""
        self._build_users()
        self._build_accounting_categories()
        self._build_businesses()
        self._build_contacts()
        self._build_jobs_and_workorders()
        self._build_purchase_orders_and_bills()
        self._backfill_default_contacts()
        self._build_tasks()
        self._build_estimates()
        self._build_invoices()
        self._build_bleps()
        self._build_price_list_items()
        self._build_implicit_jobs()
        self._build_recent_unlinked_estimates()
        self._reconcile_states()
        self._build_configuration()

    def _build_users(self):
        """Create dev_user and 4 worker users for blep assignment."""
        if self.verbose:
            print("  Building users...")

        # dev_user — needed for dev autologin and seed scripts
        dev_pk = self.get_next_pk('core.user')
        self.add_fixture('core.user', dev_pk, {
            'username': 'dev_user',
            'first_name': 'Dev',
            'last_name': 'User',
            'email': 'dev@localhost',
            'password': 'pbkdf2_sha256$1000000$szxKWr4DX4YNiiemLSRyVO$qx8Bb006xEfdlhciXd7u+f3j1QghIn+CN5C85knNzdI=',
            'is_active': True,
            'is_staff': True,
            'is_superuser': True,
        })

        workers = [
            {'first_name': 'Alex', 'last_name': 'Rivera', 'username': 'arivera'},
            {'first_name': 'Sam', 'last_name': 'Chen', 'username': 'schen'},
            {'first_name': 'Jordan', 'last_name': 'Kim', 'username': 'jkim'},
            {'first_name': 'Taylor', 'last_name': 'Brooks', 'username': 'tbrooks'},
        ]

        self.worker_user_pks = []
        for w in workers:
            pk = self.get_next_pk('core.user')
            self.worker_user_pks.append(pk)
            self.add_fixture('core.user', pk, {
                'username': w['username'],
                'first_name': w['first_name'],
                'last_name': w['last_name'],
                'email': f"{w['username']}@nealscnc.com",
                'password': '!',  # Unusable password
                'is_active': True,
                'is_staff': False,
                'is_superuser': False,
            })

        if self.verbose:
            print(f"    Created {len(self.worker_user_pks)} worker users (PKs: {self.worker_user_pks})")

    def _build_accounting_categories(self):
        """Create the four standard AccountingCategory objects."""
        if self.verbose:
            print("  Building accounting categories...")

        categories = [
            {'code': 'SVC', 'name': 'Service', 'taxable': False,
             'default_description': 'Professional services and labor'},
            {'code': 'MTL', 'name': 'Material', 'taxable': True,
             'default_description': 'Raw materials'},
            {'code': 'PRD', 'name': 'Product', 'taxable': True,
             'default_description': 'Finished products'},
            {'code': 'DLV', 'name': 'Delivery', 'taxable': False,
             'default_description': 'Shipping and delivery'},
        ]

        self.accounting_category_map = {}  # code -> pk
        for cat in categories:
            pk = self.get_next_pk('core.accountingcategory')
            self.add_fixture('core.accountingcategory', pk, {
                'code': cat['code'],
                'name': cat['name'],
                'taxable': cat['taxable'],
                'default_description': cat['default_description'],
                'is_active': True,
                'qbo_item_id': '',
                'qbo_expense_account_id': '',
            })
            self.accounting_category_map[cat['code']] = pk

        if self.verbose:
            print(f"    Created {len(categories)} accounting categories")

    def _classify_price_list_item(self, code: str, item_type: str) -> str:
        """Classify a price list item into an accounting category code.

        Returns one of: 'SVC', 'MTL', 'PRD', 'DLV'.
        """
        code_upper = (code or '').upper().strip()

        # Delivery items
        if code_upper.startswith('7 DELIVERY'):
            return 'DLV'

        # Hourly/service items
        if item_type == 'Hours':
            return 'SVC'

        # Service prefixes: boilerplate (1), setup (2), customer-provided (4), minimum (8)
        if code_upper.startswith(('1 ', '2 ', '4 ', '8 ')):
            return 'SVC'

        # Assembled products and kits
        product_prefixes = (
            'R1S-', 'R1T-', 'HIVE-', 'M67-', 'IR-', 'GLAMP', 'GLPVAN',
            'BAU', 'PM-', 'ARZABE', 'EDGE1', 'WESTERNDRILL', 'WESTERNLF',
        )
        if code_upper.startswith(product_prefixes):
            return 'PRD'

        # Everything else is raw material (sheet goods, etc.)
        return 'MTL'

    def _build_businesses(self):
        """Create Business objects from Contacts sheet."""
        if self.verbose:
            print("  Building businesses...")

        seen_orgs = set()
        ref_counter = 1

        # Check base fixtures for existing reference codes
        for fixture in self.fixture_data:
            if fixture['model'] == 'contacts.business':
                code = fixture['fields'].get('our_reference_code', '')
                if code.startswith('BUS-'):
                    try:
                        num = int(code.split('-')[1])
                        ref_counter = max(ref_counter, num + 1)
                    except (ValueError, IndexError):
                        pass

        for contact_row in self.filtered_contacts:
            org = contact_row.get('Organisation')
            if org and org not in seen_orgs:
                seen_orgs.add(org)
                pk = self.get_next_pk('contacts.business')
                self.business_map[org] = pk

                self.add_fixture('contacts.business', pk, {
                    'business_name': org,
                    'business_address': contact_row.get('Address 1', '') or '',
                    'business_phone': contact_row.get('Phone Number', '') or '',
                    'tax_exemption_number': contact_row.get('Contact VAT Number', '') or '',
                    'our_reference_code': f'BUS-{ref_counter:04d}',
                    'terms': None,
                    'default_contact': None,  # Backfilled after contacts are built
                })
                ref_counter += 1

        if self.verbose:
            print(f"    Created {len(self.business_map)} businesses")

    def _build_contacts(self):
        """Create Contact objects from Contacts sheet."""
        if self.verbose:
            print("  Building contacts...")

        for contact_row in self.filtered_contacts:
            first_name = contact_row.get('First Name', '').strip() if contact_row.get('First Name') else ''
            last_name = contact_row.get('Last Name', '').strip() if contact_row.get('Last Name') else ''
            org = contact_row.get('Organisation')

            # Handle missing names
            if org and not first_name and not last_name:
                first_name = '(unknown)'
                last_name = '(unknown)'

            # Skip if no name at all
            if not first_name and not last_name:
                continue

            pk = self.get_next_pk('contacts.contact')
            full_name = f"{first_name} {last_name}".strip()

            # Map contact — use org if available, else None
            self.contact_map[(org, full_name)] = pk

            # Get business FK
            business_fk = self.business_map.get(org) if org else None

            # Truncate phone numbers to 20 characters (max field length)
            work_phone = contact_row.get('Phone Number', '') or ''
            mobile_phone = contact_row.get('Mobile Phone Number', '') or ''
            work_number = str(work_phone)[:20] if work_phone else ''
            mobile_number = str(mobile_phone)[:20] if mobile_phone else ''

            # Ensure required fields have values
            email = contact_row.get('Email', '') or ''
            if not email:
                # Generate fake email from name
                slug = f"{first_name.lower()}.{last_name.lower()}".replace(' ', '').replace('(', '').replace(')', '')
                email = f"{slug}@example.com"

            if not work_number and not mobile_number:
                # Generate fake phone number
                work_number = f"555-{pk:04d}"

            self.add_fixture('contacts.contact', pk, {
                'first_name': first_name,
                'last_name': last_name,
                'email': email,
                'work_number': work_number,
                'mobile_number': mobile_number,
                'home_number': '',
                'addr1': contact_row.get('Address 1', '') or '',
                'addr2': contact_row.get('Address 2', '') or '',
                'addr3': contact_row.get('Address 3', '') or '',
                'city': contact_row.get('Town', '') or '',
                'municipality': contact_row.get('Region', '') or '',
                'postal_code': contact_row.get('Postcode', '') or '',
                'country_code': 'US',
                'business': business_fk,
            })

        if self.verbose:
            print(f"    Created {len(self.contact_map)} contacts")

    def _backfill_default_contacts(self):
        """Set default_contact on each Business fixture to its first Contact."""
        if self.verbose:
            print("  Setting default contacts on businesses...")

        # Build business_pk -> first contact_pk map
        first_contact_for_business = {}
        for fixture in self.fixture_data:
            if fixture['model'] == 'contacts.contact':
                biz_pk = fixture['fields'].get('business')
                if biz_pk and biz_pk not in first_contact_for_business:
                    first_contact_for_business[biz_pk] = fixture['pk']

        # Backfill default_contact on business fixtures
        updated = 0
        for fixture in self.fixture_data:
            if fixture['model'] == 'contacts.business':
                contact_pk = first_contact_for_business.get(fixture['pk'])
                if contact_pk:
                    fixture['fields']['default_contact'] = contact_pk
                    updated += 1

        if self.verbose:
            print(f"    Set default_contact on {updated} businesses")

    def _resolve_contact(
        self,
        org: str,
        contact_name: str,
        sheet: str,
        row: int,
        context: str
    ) -> Optional[int]:
        """
        Resolve contact reference, handling mismatches interactively.
        Returns contact PK or None.
        """
        if not org:
            return None

        # Try exact match first
        if (org, contact_name) in self.contact_map:
            return self.contact_map[(org, contact_name)]

        # Check if business exists
        if org not in self.business_map:
            if self.verbose:
                print(f"    Warning: Business '{org}' not found (referenced in {sheet} row {row})")
            return None

        # Find any contact for this business
        existing_contact = None
        existing_name = None
        existing_email = None

        for (map_org, map_name), contact_pk in self.contact_map.items():
            if map_org == org:
                existing_contact = contact_pk
                existing_name = map_name
                # Find email from fixture data
                for fixture in self.fixture_data:
                    if fixture['model'] == 'contacts.contact' and fixture['pk'] == contact_pk:
                        existing_email = fixture['fields'].get('email', '')
                        break
                break

        if not existing_contact:
            if self.verbose:
                print(f"    Warning: No contact found for business '{org}' (referenced in {sheet} row {row})")
            return None

        # Mismatch detected - prompt user
        decision = self.contact_handler.prompt_for_decision(
            business=org,
            existing_name=existing_name,
            existing_email=existing_email,
            referenced_name=contact_name,
            sheet=sheet,
            row=row,
            context=context
        )

        if decision == 'update':
            # Update existing contact's name
            self.contact_updates[existing_contact] = contact_name
            new_first, new_last = self._split_name(contact_name)
            # Update in fixture data
            for fixture in self.fixture_data:
                if fixture['model'] == 'contacts.contact' and fixture['pk'] == existing_contact:
                    fixture['fields']['first_name'] = new_first
                    fixture['fields']['last_name'] = new_last
                    break
            # Update map
            self.contact_map[(org, contact_name)] = existing_contact
            return existing_contact

        elif decision == 'create':
            # Create new contact
            pk = self.get_next_pk('contacts.contact')
            new_first, new_last = self._split_name(contact_name)

            # Generate email/phone
            slug = f"{new_first.lower()}.{new_last.lower()}".replace(' ', '')
            new_email = f"{slug}@example.com"

            self.add_fixture('contacts.contact', pk, {
                'first_name': new_first,
                'last_name': new_last,
                'email': new_email,
                'work_number': f"555-{pk:04d}",
                'mobile_number': '',
                'home_number': '',
                'addr1': '',
                'addr2': '',
                'addr3': '',
                'city': '',
                'municipality': '',
                'postal_code': '',
                'country_code': 'US',
                'business': self.business_map[org],
            })

            self.contact_map[(org, contact_name)] = pk
            return pk

        else:  # 'map'
            # Use existing contact as-is
            self.contact_map[(org, contact_name)] = existing_contact
            return existing_contact

    def _get_v1_estimate_dates(self) -> Dict[str, datetime]:
        """
        Pre-scan estimates to find V1 estimate dates for each project.
        Returns a map of project_name -> date of V1 estimate.
        """
        v1_dates = {}

        for estimate in self.filtered_estimates:
            project_name = estimate.get('Project')
            if not project_name:
                continue

            # Parse revision from reference
            reference = estimate.get('Reference', '') or ''
            base_ref, revision = self._parse_revision_suffix(reference)

            # Only interested in V1 estimates
            if revision == 1:
                est_date = estimate.get('Date')
                if est_date and isinstance(est_date, datetime):
                    # If we already have a V1 date for this project, use the earliest one
                    if project_name not in v1_dates or est_date < v1_dates[project_name]:
                        v1_dates[project_name] = est_date

        return v1_dates

    def _has_estimates(self, project_name: str) -> bool:
        """Check if a project has any estimates."""
        for estimate in self.filtered_estimates:
            if estimate.get('Project') == project_name:
                return True
        return False

    def _build_jobs_and_workorders(self):
        """Create Job and WorkOrder objects from Projects sheet."""
        if self.verbose:
            print("  Building jobs and work orders...")

        status_map = {
            'Completed': 'completed',
            'Active': 'approved',
            'Cancelled': 'cancelled',
        }

        # Pre-scan estimates to get V1 dates for start_date calculation
        v1_estimate_dates = self._get_v1_estimate_dates()

        # Track job number counters per year
        job_counters = {}

        for project in self.filtered_projects:
            project_name = project.get('Name')
            if not project_name:
                continue

            # Create Job
            job_pk = self.get_next_pk('jobs.job')

            # Resolve contact - required for Job model
            client_org = project.get('Client Organisation')
            client_name = project.get('Client Name')
            contact_fk = None

            # Try normal resolution first (org + name)
            if client_org and client_name:
                contact_fk = self._resolve_contact(
                    org=client_org,
                    contact_name=client_name,
                    sheet='Projects',
                    row=project.get('_row', 0),
                    context=project_name
                )

            # If no org or normal resolution failed, try to find contact by name only
            if not contact_fk and client_name:
                for (map_org, map_name), pk in self.contact_map.items():
                    if map_name == client_name:
                        contact_fk = pk
                        if self.verbose:
                            print(f"    Found contact '{client_name}' by name match (org='{map_org}')")
                        break

            # If org but no name, find any contact for that business
            if not contact_fk and client_org and client_org in self.business_map:
                biz_pk = self.business_map[client_org]
                for (map_org, map_name), pk in self.contact_map.items():
                    if map_org == client_org:
                        contact_fk = pk
                        if self.verbose:
                            print(f"    Found contact '{map_name}' for org '{client_org}' (no client name given)")
                        break

            # Skip this job if we still couldn't resolve a contact (Job model requires contact)
            if not contact_fk:
                if self.verbose:
                    print(f"    Skipping job '{project_name}' - no valid contact (org='{client_org}', name='{client_name}')")
                continue

            business_fk = self.business_map.get(client_org) if client_org else None

            job_status = status_map.get(project.get('Status'), 'active')

            # Calculate dates based on new rules:
            # created_date: Created Date from spreadsheet
            created_date = self._format_date(project.get('Created Date'))

            # start_date: If project has explicit "Starts On", use it;
            #             otherwise if approved, use V1 estimate date;
            #             otherwise if completed and no estimates, use created_date
            start_date = self._format_date(project.get('Starts On'))
            if not start_date and job_status == 'approved':
                # Use V1 estimate date if available
                v1_date = v1_estimate_dates.get(project_name)
                if v1_date:
                    start_date = self._format_date(v1_date)

            if not start_date and job_status == 'completed':
                # If no estimates exist for this project, use created_date
                if not self._has_estimates(project_name):
                    start_date = created_date

            # due_date: If project has explicit "Ends On", use it; otherwise leave blank
            due_date = self._format_date(project.get('Ends On'))

            # completed_date: If approved (Active), leave blank; otherwise use Updated Date
            if job_status == 'approved':
                completed_date = None
            else:
                completed_date = self._format_date(project.get('Updated Date'))

            # Generate job number in format J{year}-{counter:04d}
            created_dt = project.get('Created Date')
            if isinstance(created_dt, datetime):
                year = created_dt.year
            else:
                year = 2025  # Default year if no date

            # Increment counter for this year
            if year not in job_counters:
                job_counters[year] = 1
            else:
                job_counters[year] += 1

            job_number = f"J{year}-{job_counters[year]:04d}"

            self.add_fixture('jobs.job', job_pk, {
                'name': project_name[:50],
                'job_number': job_number,
                'contact': contact_fk,
                'start_date': start_date,
                'due_date': due_date,
                'created_date': created_date,
                'customer_po_number': project.get('Contract PO Reference', '') or '',
                'status': job_status,
                'description': project.get('Notes', '') or '',
                'completed_date': completed_date,
            })

            # Create WorkOrder if not cancelled
            workorder_pk = None
            if project.get('Status') != 'CANCELLED':
                workorder_pk = self.get_next_pk('jobs.workorder')

                wo_status = 'complete' if job_status == 'completed' else 'incomplete'

                self.add_fixture('jobs.workorder', workorder_pk, {
                    'job': job_pk,
                    'status': wo_status,
                    'template': None,
                })

            self.job_map[project_name] = (job_pk, workorder_pk)

        if self.verbose:
            jobs_count = len(self.job_map)
            wo_count = sum(1 for _, wo_pk in self.job_map.values() if wo_pk is not None)
            print(f"    Created {jobs_count} jobs and {wo_count} work orders")

    def _build_purchase_orders_and_bills(self):
        """Create PurchaseOrder and Bill objects from Bills sheet."""
        if self.verbose:
            print("  Building purchase orders and bills...")

        # Track bill number counters per year
        bill_counters = {}
        po_counters = {}

        for bill in self.filtered_bills:
            line_items = bill.get('_line_items', [])
            self._save_bill_and_po(bill, line_items, bill_counters, po_counters)

        if self.verbose:
            print(f"    Created {self.pk_counters.get('purchasing.purchaseorder', 1) - 1} purchase orders")
            print(f"    Created {self.pk_counters.get('purchasing.bill', 1) - 1} bills")

    def _ensure_individual_vendor(self, contact_name: str) -> Tuple[int, int]:
        """Create a personal Business + Contact for an individual vendor (no org in source data).
        Returns (business_pk, contact_pk)."""
        # Use the person's name as a synthetic org key
        synthetic_org = f"(individual) {contact_name}"

        if synthetic_org in self.business_map:
            # Already created — find the contact
            contact_pk = self.contact_map.get((synthetic_org, contact_name))
            return self.business_map[synthetic_org], contact_pk

        # Create contact first (we need the PK for default_contact)
        contact_pk = self.get_next_pk('contacts.contact')
        first_name, last_name = self._split_name(contact_name)
        slug = f"{first_name.lower()}.{last_name.lower()}".replace(' ', '')

        self.add_fixture('contacts.contact', contact_pk, {
            'first_name': first_name,
            'last_name': last_name,
            'email': f"{slug}@example.com",
            'work_number': f"555-{contact_pk:04d}",
            'mobile_number': '',
            'home_number': '',
            'addr1': '',
            'addr2': '',
            'addr3': '',
            'city': '',
            'municipality': '',
            'postal_code': '',
            'country_code': 'US',
            'business': None,  # Will be set below after business is created
        })

        # Create business
        biz_pk = self.get_next_pk('contacts.business')
        # Generate reference code
        ref_counter = 1
        for fixture in self.fixture_data:
            if fixture['model'] == 'contacts.business':
                code = fixture['fields'].get('our_reference_code', '')
                if code.startswith('BUS-'):
                    try:
                        num = int(code.split('-')[1])
                        ref_counter = max(ref_counter, num + 1)
                    except (ValueError, IndexError):
                        pass

        self.add_fixture('contacts.business', biz_pk, {
            'business_name': contact_name,
            'business_address': '',
            'business_phone': '',
            'tax_exemption_number': '',
            'our_reference_code': f'BUS-{ref_counter:04d}',
            'terms': None,
            'default_contact': contact_pk,
        })

        # Update contact's business FK
        for fixture in self.fixture_data:
            if fixture['model'] == 'contacts.contact' and fixture['pk'] == contact_pk:
                fixture['fields']['business'] = biz_pk
                break

        self.business_map[synthetic_org] = biz_pk
        self.contact_map[(synthetic_org, contact_name)] = contact_pk

        if self.verbose:
            print(f"    Created individual vendor: {contact_name} (business PK {biz_pk})")

        return biz_pk, contact_pk

    def _save_bill_and_po(self, bill_row: Dict, line_items: List[Dict], bill_counters: Dict, po_counters: Dict):
        """Save a Bill and its associated PurchaseOrder and line items."""
        # Create PurchaseOrder
        po_pk = self.get_next_pk('purchasing.purchaseorder')

        # Resolve contact and business
        contact_org = bill_row.get('Contact Organisation')
        contact_name = bill_row.get('Contact Name')
        contact_fk = None
        business_fk = None

        if contact_org:
            # Normal case: has an organisation
            if contact_name:
                contact_fk = self._resolve_contact(
                    org=contact_org,
                    contact_name=contact_name,
                    sheet='Bills',
                    row=bill_row.get('_row', 0),
                    context=bill_row.get('Reference', 'Unknown')
                )
            business_fk = self.business_map.get(contact_org)
        elif contact_name:
            # Individual vendor: no org, just a person's name
            business_fk, contact_fk = self._ensure_individual_vendor(contact_name)
        else:
            # Neither org nor name — skip this bill
            if self.verbose:
                print(f"    Warning: Skipping bill row {bill_row.get('_row', '?')} — no contact info")
            return

        # Get job FK
        project_name = bill_row.get('Project')
        job_fk = None
        if project_name and project_name in self.job_map:
            job_fk, _ = self.job_map[project_name]

        # Dates
        created_date = self._format_date(bill_row.get('Date'))

        # Generate PO number in format PO{year}-{counter:04d}
        created_dt = bill_row.get('Date')
        if isinstance(created_dt, datetime):
            year = created_dt.year
        else:
            year = 2025  # Default year if no date

        # Increment PO counter for this year
        if year not in po_counters:
            po_counters[year] = 1
        else:
            po_counters[year] += 1

        po_number = f"PO{year}-{po_counters[year]:04d}"

        # Line item requirement: cannot leave draft without line items
        if line_items:
            po_status = 'issued'
            po_issued_date = created_date
        else:
            if self.verbose:
                print(f"    Warning: PO {po_number} has no line items, keeping as draft")
            po_status = 'draft'
            po_issued_date = None

        self.add_fixture('purchasing.purchaseorder', po_pk, {
            'po_number': po_number,
            'business': business_fk,
            'contact': contact_fk,
            'status': po_status,
            'created_date': created_date,
            'requested_date': None,
            'issued_date': po_issued_date,
            'received_date': None,
            'cancel_date': None,
        })

        # Create Bill
        bill_pk = self.get_next_pk('purchasing.bill')

        due_date = self._format_date(bill_row.get('Due Date'))
        reference = bill_row.get('Reference', '')

        # Generate Bill number in format B{year}-{counter:04d}
        if year not in bill_counters:
            bill_counters[year] = 1
        else:
            bill_counters[year] += 1

        bill_number = f"B{year}-{bill_counters[year]:04d}"

        self.add_fixture('purchasing.bill', bill_pk, {
            'bill_number': bill_number,
            'purchase_order': po_pk,
            'business': business_fk,
            'contact': contact_fk,
            'vendor_invoice_number': reference or '',
            'status': 'draft',
            'created_date': created_date,
            'due_date': due_date,
            'received_date': None,
            'paid_date': None,
            'cancelled_date': None,
        })

        # Create line items for both PO and Bill with sequential line numbers starting at 1
        line_number = 1
        for item_row in line_items:
            # PO Line Item
            po_item_pk = self.get_next_pk('purchasing.purchaseorderlineitem')

            qty = self._parse_decimal(item_row.get('Quantity', 1)) or Decimal('1')
            net_value = abs(self._parse_decimal(item_row.get('Net Value', 0)))
            price = net_value / qty

            self.add_fixture('purchasing.purchaseorderlineitem', po_item_pk, {
                'purchase_order': po_pk,
                'job': job_fk,
                'task': None,
                'price_list_item': None,
                'line_number': line_number,
                'qty': str(qty),
                'units': item_row.get('Item Type', '-no unit-') or '-no unit-',
                'description': item_row.get('Description', '') or '',
                'price': str(price),
            })

            # Bill Line Item
            bill_item_pk = self.get_next_pk('purchasing.billlineitem')

            self.add_fixture('purchasing.billlineitem', bill_item_pk, {
                'bill': bill_pk,
                'task': None,
                'price_list_item': None,
                'line_number': line_number,
                'qty': str(qty),
                'units': item_row.get('Item Type', '-no unit-') or '-no unit-',
                'description': item_row.get('Description', '') or '',
                'price': str(price),
            })

            line_number += 1

    def _build_tasks(self):
        """Create Task objects from Tasks sheet."""
        if self.verbose:
            print("  Building tasks...")

        task_status_map = {
            'Completed': 'complete',
            'Active': 'in_progress',
        }

        # Build set of (project, task) pairs that have timeslips
        tasks_with_timeslips = set()
        for ts in self.filtered_timeslips:
            proj = ts.get('Project')
            task = ts.get('Task')
            if proj and task:
                tasks_with_timeslips.add((proj, task))

        # Build WO PK -> WO status lookup
        wo_status_map = {}
        for fixture in self.fixture_data:
            if fixture['model'] == 'jobs.workorder':
                wo_status_map[fixture['pk']] = fixture['fields']['status']

        # Group tasks by project/work order
        tasks_by_wo = {}

        for task_row in self.filtered_tasks:
            project_name = task_row.get('Project')
            if not project_name or project_name not in self.job_map:
                continue

            _, workorder_pk = self.job_map[project_name]
            if not workorder_pk:
                continue  # Skip if no work order (cancelled project)

            if workorder_pk not in tasks_by_wo:
                tasks_by_wo[workorder_pk] = []

            tasks_by_wo[workorder_pk].append(task_row)

        # Create tasks with line numbers
        for workorder_pk, tasks in tasks_by_wo.items():
            wo_status = wo_status_map.get(workorder_pk, 'incomplete')

            for line_num, task_row in enumerate(tasks, start=1):
                task_pk = self.get_next_pk('jobs.task')

                task_name = task_row.get('Name', '')
                project_name = task_row.get('Project', '')

                # Store in task map for Blep lookup
                self.task_map[(project_name, task_name)] = task_pk

                rate = self._parse_decimal(task_row.get('Billing Rate', 0))

                # Determine task status:
                # - Tasks on complete WOs must be complete
                # - Source 'Completed' → complete
                # - Source 'Active' with timeslips → in_progress
                # - Source 'Active' without timeslips → pending
                source_status = task_row.get('Status', '')
                if wo_status == 'complete':
                    task_status = 'complete'
                elif source_status in task_status_map:
                    task_status = task_status_map[source_status]
                else:
                    task_status = 'pending'

                # Tasks with timeslips must be at least in_progress
                has_timeslips = (project_name, task_name) in tasks_with_timeslips
                if has_timeslips and task_status == 'pending':
                    task_status = 'in_progress'

                self.add_fixture('jobs.task', task_pk, {
                    'parent_task': None,
                    'assignee': None,
                    'work_order': workorder_pk,
                    'est_worksheet': None,
                    'name': task_name,
                    'sort_order': line_num,
                    'status': task_status,
                    'units': 'hours',
                    'rate': str(rate),
                    'est_qty': '0',
                })

        if self.verbose:
            print(f"    Created {len(self.task_map)} tasks")

    def _create_additional_job(self, original_project_name: str, estimate_base_ref: str, part_number: int,
                                job_counters: Dict, v1_estimate_dates: Dict) -> Tuple[int, int]:
        """Create an additional job for a multi-estimate project."""
        # Find the original project data
        original_project = None
        for project in self.filtered_projects:
            if project.get('Name') == original_project_name:
                original_project = project
                break

        if not original_project:
            return None, None

        # Get original job info
        original_job_fk, _ = self.job_map.get(original_project_name, (None, None))

        # Create new job
        job_pk = self.get_next_pk('jobs.job')

        # Get contact info (same as original)
        client_org = original_project.get('Client Organisation')
        client_name = original_project.get('Client Name')
        contact_fk = None

        if client_org and client_name:
            contact_fk = self._resolve_contact(
                org=client_org,
                contact_name=client_name,
                sheet='Projects',
                row=original_project.get('_row', 0),
                context=original_project_name
            )

        if not contact_fk and client_name:
            for (org, name), pk in self.contact_map.items():
                if name == client_name:
                    contact_fk = pk
                    break

        if not contact_fk:
            return None, None

        business_fk = self.business_map.get(client_org) if client_org else None

        # Generate job number
        created_dt = original_project.get('Created Date')
        if isinstance(created_dt, datetime):
            year = created_dt.year
        else:
            year = 2025

        if year not in job_counters:
            job_counters[year] = 1
        else:
            job_counters[year] += 1

        job_number = f"J{year}-{job_counters[year]:04d}"

        # Create modified job name (max 50 chars for Job.name field)
        # Format: "Original Name - Est XXXX"
        suffix = f" - Est {estimate_base_ref}"
        max_base_len = 50 - len(suffix)

        if len(original_project_name) <= max_base_len:
            new_job_name = original_project_name + suffix
        else:
            # Truncate original name to fit
            new_job_name = original_project_name[:max_base_len] + suffix

        # Get dates and status
        job_status = original_project.get('Status', 'Active')
        status_map = {'Completed': 'completed', 'Active': 'approved', 'Cancelled': 'cancelled'}
        job_status = status_map.get(job_status, 'approved')

        # Calculate dates using same rules as main job creation
        created_date = self._format_date(original_project.get('Created Date'))

        # start_date: If project has explicit "Starts On", use it;
        #             otherwise if approved, use V1 estimate date;
        #             otherwise if completed and no estimates, use created_date
        start_date = self._format_date(original_project.get('Starts On'))
        if not start_date and job_status == 'approved':
            # Use V1 estimate date if available
            v1_date = v1_estimate_dates.get(original_project_name)
            if v1_date:
                start_date = self._format_date(v1_date)

        if not start_date and job_status == 'completed':
            # If no estimates exist for this project, use created_date
            if not self._has_estimates(original_project_name):
                start_date = created_date

        # due_date: If project has explicit "Ends On", use it; otherwise leave blank
        due_date = self._format_date(original_project.get('Ends On'))

        # completed_date: If approved (Active), leave blank; otherwise use Updated Date
        if job_status == 'approved':
            completed_date = None
        else:
            completed_date = self._format_date(original_project.get('Updated Date'))

        # Add reference to original job in description
        original_job_num = None
        for name, (_, _) in self.job_map.items():
            if name == original_project_name:
                # Find the original job number from fixtures
                for fixture in self.fixture_data:
                    if fixture['model'] == 'jobs.job' and fixture['fields'].get('name') == original_project_name:
                        original_job_num = fixture['fields'].get('job_number')
                        break
                break

        description = original_project.get('Notes', '') or ''
        if original_job_num:
            description = f"Related to Job {original_job_num}. " + description

        self.add_fixture('jobs.job', job_pk, {
            'name': new_job_name,
            'job_number': job_number,
            'contact': contact_fk,
            'start_date': start_date,
            'due_date': due_date,
            'created_date': created_date,
            'customer_po_number': original_project.get('Contract PO Reference', '') or '',
            'status': job_status,
            'description': description,
            'completed_date': completed_date,
        })

        # Create workorder
        workorder_pk = self.get_next_pk('jobs.workorder')
        wo_status = 'complete' if job_status == 'completed' else 'incomplete'

        self.add_fixture('jobs.workorder', workorder_pk, {
            'job': job_pk,
            'status': wo_status,
            'template': None,
        })

        return job_pk, workorder_pk

    def _build_estimates(self):
        """Create Estimate and EstimateLineItem objects."""
        if self.verbose:
            print("  Building estimates...")

        status_map = {
            'Draft': 'draft',
            'Sent': 'open',
            'Approved': 'accepted',
            'Rejected': 'rejected',
        }

        # Get V1 estimate dates for job date calculations
        v1_estimate_dates = self._get_v1_estimate_dates()

        # First, group estimates by project to detect multi-estimate situations
        estimates_by_project = {}
        for estimate in self.filtered_estimates:
            project_name = estimate.get('Project')
            if not project_name or project_name not in self.job_map:
                continue

            if project_name not in estimates_by_project:
                estimates_by_project[project_name] = []

            estimates_by_project[project_name].append(estimate)

        # Track estimate PKs by base reference and revision for parent relationships
        estimate_pk_map = {}
        estimate_count = 0
        line_item_count = 0
        additional_jobs_created = 0

        # Job counters for creating additional jobs
        job_counters = {}

        # Initialize counters from existing jobs
        for fixture in self.fixture_data:
            if fixture['model'] == 'jobs.job':
                job_num = fixture['fields'].get('job_number', '')
                if job_num.startswith('J'):
                    try:
                        parts = job_num.split('-')
                        if len(parts) == 2:
                            year = int(parts[0][1:])
                            counter = int(parts[1])
                            if year not in job_counters or counter > job_counters[year]:
                                job_counters[year] = counter
                    except:
                        pass

        # Process each project's estimates
        for project_name, project_estimates in estimates_by_project.items():
            # Group by base reference number within this project
            estimates_by_base_ref = {}

            for estimate in project_estimates:
                reference = estimate.get('Reference', '') or ''
                base_ref, revision = self._parse_revision_suffix(reference)

                if base_ref not in estimates_by_base_ref:
                    estimates_by_base_ref[base_ref] = []

                estimates_by_base_ref[base_ref].append({
                    'estimate': estimate,
                    'base_ref': base_ref,
                    'revision': revision,
                    'reference': reference,
                })

            # Sort estimates by date to determine which is "first"
            base_refs_sorted = sorted(
                estimates_by_base_ref.keys(),
                key=lambda br: min(
                    e['estimate'].get('Date') or datetime(1900, 1, 1)
                    for e in estimates_by_base_ref[br]
                )
            )

            # Track jobs for this project's estimates
            jobs_for_estimates = {}

            # First base ref uses the original job
            first_base_ref = base_refs_sorted[0]
            original_job_fk, original_wo_fk = self.job_map[project_name]
            jobs_for_estimates[first_base_ref] = (original_job_fk, original_wo_fk)

            # Additional base refs need new jobs
            for i, base_ref in enumerate(base_refs_sorted[1:], start=2):
                new_job_fk, new_wo_fk = self._create_additional_job(
                    project_name, base_ref, i, job_counters, v1_estimate_dates
                )
                if new_job_fk:
                    jobs_for_estimates[base_ref] = (new_job_fk, new_wo_fk)
                    additional_jobs_created += 1
                    if self.verbose:
                        print(f"    Created additional job for {project_name} / estimate {base_ref}")

            # Now create estimates, linking to appropriate jobs
            for base_ref in base_refs_sorted:
                if base_ref not in jobs_for_estimates:
                    continue

                job_fk, _ = jobs_for_estimates[base_ref]
                revisions = estimates_by_base_ref[base_ref]

                # Sort revisions
                revisions.sort(key=lambda x: x['revision'])
                max_revision = max(r['revision'] for r in revisions)

                for rev_data in revisions:
                    estimate = rev_data['estimate']
                    revision = rev_data['revision']

                    # Create Estimate
                    estimate_pk = self.get_next_pk('estimates.estimate')
                    estimate_count += 1

                    # Store PK for parent relationship
                    estimate_pk_map[(base_ref, revision)] = estimate_pk

                    # Determine parent FK
                    parent_fk = None
                    if revision > 1:
                        parent_fk = estimate_pk_map.get((base_ref, revision - 1))

                    # Determine status
                    status = status_map.get(estimate.get('Status'), 'draft')
                    if revision < max_revision:
                        status = 'superseded'

                    # Line item requirement: cannot leave draft without line items
                    # (superseded is exempt — set by version chain, not by transition)
                    line_items = estimate.get('_line_items', [])
                    if status not in ('draft', 'superseded') and not line_items:
                        if self.verbose:
                            print(f"    Warning: Estimate {reference} has no line items, keeping as draft")
                        status = 'draft'

                    created_date = self._format_date(estimate.get('Date'))

                    self.add_fixture('estimates.estimate', estimate_pk, {
                        'job': job_fk,
                        'estimate_number': base_ref,
                        'version': revision,
                        'status': status,
                        'parent': parent_fk,
                        'created_date': created_date,
                        'sent_date': created_date if status in ['open', 'accepted', 'rejected'] else None,
                        'closed_date': None,
                        'expiration_date': None,
                    })

                    # Create line items with sequential line numbers starting at 1
                    line_number = 1
                    for item_row in estimate.get('_line_items', []):
                        line_item_pk = self.get_next_pk('estimates.estimatelineitem')
                        line_item_count += 1

                        qty = self._parse_decimal(item_row.get('Quantity', 1))
                        price = self._parse_decimal(item_row.get('Price', 0))

                        self.add_fixture('estimates.estimatelineitem', line_item_pk, {
                            'estimate': estimate_pk,
                            'task': None,
                            'price_list_item': None,
                            'line_number': line_number,
                            'qty': str(qty),
                            'units': '',
                            'description': item_row.get('Description', '') or '',
                            'price': str(price),
                        })

                        line_number += 1

        if self.verbose:
            print(f"    Created {estimate_count} estimates with {line_item_count} line items")
            if additional_jobs_created > 0:
                print(f"    Created {additional_jobs_created} additional jobs for multi-estimate projects")

    def _build_invoices(self):
        """Create Invoice and InvoiceLineItem objects."""
        if self.verbose:
            print("  Building invoices...")

        status_map = {
            'Draft': 'draft',
            'Sent': 'open',
            'Paid': 'paid',
            'Part Paid': 'partly-paid',
            'Overdue': 'open',
            'Cancelled': 'cancelled',
        }

        invoice_count = 0
        line_item_count = 0

        for invoice in self.filtered_invoices:
            # Get job FK from Projects field
            projects_str = invoice.get('Projects', '')
            job_fk = None

            if projects_str:
                # Try first project name in the comma-separated list
                for project_name in projects_str.split(','):
                    project_name = project_name.strip()
                    if project_name in self.job_map:
                        job_fk, _ = self.job_map[project_name]
                        break

            if not job_fk:
                continue

            # Get business FK
            contact_org = invoice.get('Contact Organisation')
            business_fk = self.business_map.get(contact_org) if contact_org else None

            # Create Invoice
            invoice_pk = self.get_next_pk('invoicing.invoice')
            invoice_count += 1

            status = status_map.get(invoice.get('Status'), 'draft')

            # FreeAgent marks paid invoices as 'Sent' with a Paid Date
            paid_date = invoice.get('Paid Date')
            if status == 'open' and paid_date:
                paid_amount = float(invoice.get('Paid Amount', 0) or 0)
                total_value = float(invoice.get('Total Value', 0) or 0)
                if total_value and abs(paid_amount - total_value) < 0.01:
                    status = 'paid'

            # Line item requirement: cannot leave draft without line items
            if status != 'draft' and not invoice.get('_line_items'):
                if self.verbose:
                    print(f"    Warning: Invoice {invoice.get('Reference', '?')} has no line items, keeping as draft")
                status = 'draft'

            created_date = self._format_date(invoice.get('Date'))
            closed_date = self._format_date(paid_date) if status == 'paid' else None

            self.add_fixture('invoicing.invoice', invoice_pk, {
                'job': job_fk,
                'invoice_number': invoice.get('Reference', '') or '',
                'status': status,
                'created_date': created_date,
                'sent_date': created_date if status in ['open', 'paid', 'partly-paid'] else None,
                'closed_date': closed_date,
            })

            # Create line items with sequential line numbers starting at 1
            line_number = 1
            for item_row in invoice.get('_line_items', []):
                line_item_pk = self.get_next_pk('invoicing.invoicelineitem')
                line_item_count += 1

                qty = self._parse_decimal(item_row.get('Quantity', 1))
                price = self._parse_decimal(item_row.get('Price', 0))

                self.add_fixture('invoicing.invoicelineitem', line_item_pk, {
                    'invoice': invoice_pk,
                    'task': None,
                    'price_list_item': None,
                    'line_number': line_number,
                    'qty': str(qty),
                    'units': '',
                    'description': item_row.get('Description', '') or '',
                    'price': str(price),
                })

                line_number += 1

        if self.verbose:
            print(f"    Created {invoice_count} invoices with {line_item_count} line items")

    def _build_bleps(self):
        """Create Blep objects from Timeslips sheet."""
        if self.verbose:
            print("  Building bleps (time tracking)...")

        blep_user_idx = 0  # Round-robin index

        for timeslip in self.filtered_timeslips:
            # Find task
            project_name = timeslip.get('Project')
            task_name = timeslip.get('Task')

            if not project_name or not task_name:
                continue

            task_pk = self.task_map.get((project_name, task_name))
            if not task_pk:
                if self.verbose:
                    print(f"    Warning: Task '{task_name}' not found for project '{project_name}'")
                continue

            # Calculate start and end times
            date = timeslip.get('Date')
            hours = timeslip.get('Hours', 0)

            if not date or not isinstance(date, datetime):
                continue

            # Start time: 9:00 AM on the date
            start_time = datetime.combine(date.date(), datetime.min.time().replace(hour=9))
            # End time: start + hours
            end_time = start_time + timedelta(hours=float(hours) if hours else 0)

            blep_pk = self.get_next_pk('jobs.blep')

            # Assign user round-robin among worker users
            user_pk = self.worker_user_pks[blep_user_idx % len(self.worker_user_pks)]
            blep_user_idx += 1

            self.add_fixture('jobs.blep', blep_pk, {
                'user': user_pk,
                'task': task_pk,
                'start_time': start_time.isoformat(),
                'end_time': end_time.isoformat(),
            })

        if self.verbose:
            blep_count = self.pk_counters.get('jobs.blep', 1) - 1
            print(f"    Created {blep_count} bleps")

    def _build_price_list_items(self):
        """Create PriceListItem objects."""
        if self.verbose:
            print("  Building price list items...")

        category_counts = {}
        for item in self.filtered_price_list:
            pk = self.get_next_pk('inventory.pricelistitem')

            qty = self._parse_decimal(item.get('Quantity', 1))
            price = self._parse_decimal(item.get('Price', 0))

            code = item.get('Code', '') or ''
            item_type = item.get('Type', '') or ''
            cat_code = self._classify_price_list_item(code, item_type)
            cat_pk = self.accounting_category_map[cat_code]
            category_counts[cat_code] = category_counts.get(cat_code, 0) + 1

            self.add_fixture('inventory.pricelistitem', pk, {
                'code': code,
                'units': item_type,
                'description': item.get('Description', '') or '',
                'purchase_price': str(price),
                'selling_price': str(price),
                'qty_on_hand': str(qty),
                'qty_sold': '0',
                'qty_wasted': '0',
                'is_active': True,
                'accounting_category': cat_pk,
            })

        if self.verbose:
            print(f"    Created {len(self.filtered_price_list)} price list items")
            for cat_code, count in sorted(category_counts.items()):
                print(f"      {cat_code}: {count}")

    def _resolve_or_create_contact_for_estimate(self, estimate: Dict) -> Tuple[Optional[int], Optional[int]]:
        """Resolve or create contact + business for an implicit job from estimate data.
        Returns (contact_pk, business_pk)."""
        org = estimate.get('Sent to Contact Organisation') or ''
        first_name = estimate.get('Sent to Contact First Name', '') or ''
        last_name = estimate.get('Sent to Contact Last Name', '') or ''
        full_name = f"{first_name} {last_name}".strip()

        if not full_name and not org:
            # Try the Contact field as a fallback
            contact_field = estimate.get('Contact', '') or ''
            if contact_field:
                full_name = contact_field
                first_name, last_name = self._split_name(contact_field)

        if not full_name:
            return None, None

        # Try to find existing contact
        if org and (org, full_name) in self.contact_map:
            biz_pk = self.business_map.get(org)
            return self.contact_map[(org, full_name)], biz_pk

        # Try name-only match
        for (map_org, map_name), pk in self.contact_map.items():
            if map_name == full_name:
                biz_pk = self.business_map.get(map_org) if map_org else None
                return pk, biz_pk

        # Create new business + contact
        if org and org in self.business_map:
            biz_pk = self.business_map[org]
        elif org:
            biz_pk = self.get_next_pk('contacts.business')
            ref_counter = sum(1 for f in self.fixture_data if f['model'] == 'contacts.business') + 1
            self.add_fixture('contacts.business', biz_pk, {
                'business_name': org,
                'business_address': '',
                'business_phone': '',
                'tax_exemption_number': '',
                'our_reference_code': f'BUS-{ref_counter:04d}',
                'terms': None,
                'default_contact': None,
            })
            self.business_map[org] = biz_pk
        else:
            # Individual — create personal business
            biz_pk, contact_pk = self._ensure_individual_vendor(full_name)
            return contact_pk, biz_pk

        # Create contact
        if not first_name:
            first_name, last_name = self._split_name(full_name)

        contact_pk = self.get_next_pk('contacts.contact')
        slug = f"{first_name.lower()}.{last_name.lower()}".replace(' ', '').replace('(', '').replace(')', '')

        self.add_fixture('contacts.contact', contact_pk, {
            'first_name': first_name,
            'last_name': last_name,
            'email': f"{slug}@example.com",
            'work_number': f"555-{contact_pk:04d}",
            'mobile_number': '',
            'home_number': '',
            'addr1': '',
            'addr2': '',
            'addr3': '',
            'city': '',
            'municipality': '',
            'postal_code': '',
            'country_code': 'US',
            'business': biz_pk,
        })
        self.contact_map[(org or None, full_name)] = contact_pk

        # Set default_contact on business if not yet set
        for f in self.fixture_data:
            if f['model'] == 'contacts.business' and f['pk'] == biz_pk:
                if not f['fields'].get('default_contact'):
                    f['fields']['default_contact'] = contact_pk
                break

        return contact_pk, biz_pk

    def _create_implicit_job(
        self, estimate: Dict, invoices: List[Dict], bills: List[Dict],
        job_counters: Dict, po_counters: Dict, bill_counters: Dict
    ) -> Optional[int]:
        """Create a full implicit Job from an unlinked estimate + its linked invoices/bills.
        Returns the job PK or None."""
        import re

        est_ref = estimate.get('Reference', '') or ''
        est_date = estimate.get('Date')

        # Resolve contact
        contact_pk, business_pk = self._resolve_or_create_contact_for_estimate(estimate)
        if not contact_pk:
            return None

        # Determine job status from estimate + invoice states
        est_status_map = {
            'Draft': 'draft', 'Sent': 'open',
            'Approved': 'accepted', 'Rejected': 'rejected',
        }
        est_status = est_status_map.get(estimate.get('Status'), 'draft')

        inv_status_map = {
            'Draft': 'draft', 'Sent': 'open',
            'Cancelled': 'cancelled',
        }

        # Determine if all invoices are paid
        all_paid = bool(invoices)
        for inv in invoices:
            inv_st = inv_status_map.get(inv.get('Status'), 'draft')
            paid_date = inv.get('Paid Date')
            if inv_st == 'open' and paid_date:
                paid_amount = float(inv.get('Paid Amount', 0) or 0)
                total_value = float(inv.get('Total Value', 0) or 0)
                if total_value and abs(paid_amount - total_value) < 0.01:
                    continue  # This one is paid
            all_paid = False

        if all_paid:
            job_status = 'completed'
        elif est_status == 'accepted':
            job_status = 'approved'
        elif est_status in ('open', 'rejected'):
            job_status = 'submitted'
        else:
            job_status = 'draft'

        # Dates
        created_date = self._format_date(est_date)
        year = est_date.year if isinstance(est_date, datetime) else 2025

        if year not in job_counters:
            job_counters[year] = 0
        job_counters[year] += 1
        job_number = f"J{year}-{job_counters[year]:04d}"

        start_date = created_date if job_status in ('approved', 'completed') else None
        completed_date = None
        if job_status == 'completed' and invoices:
            # Use latest paid date
            paid_dates = [self._format_date(i.get('Paid Date')) for i in invoices if i.get('Paid Date')]
            completed_date = max(paid_dates) if paid_dates else created_date

        # Job name from estimate reference or description
        est_line_items = estimate.get('_line_items', [])
        # Use second line item for name (first is usually boilerplate/setup)
        desc_item = est_line_items[1] if len(est_line_items) > 1 else est_line_items[0] if est_line_items else None
        first_desc = (desc_item.get('Description', '') or '')[:40] if desc_item else ''
        job_name = f"Est {est_ref}"
        if first_desc:
            job_name = f"Est {est_ref} - {first_desc}"
        job_name = job_name[:50]

        # Create Job
        job_pk = self.get_next_pk('jobs.job')

        self.add_fixture('jobs.job', job_pk, {
            'name': job_name,
            'job_number': job_number,
            'contact': contact_pk,
            'start_date': start_date,
            'due_date': None,
            'created_date': created_date,
            'customer_po_number': '',
            'status': job_status,
            'description': estimate.get('Notes', '') or '',
            'completed_date': completed_date,
        })

        # Create Estimate + line items
        estimate_pk = self.get_next_pk('estimates.estimate')

        # Line item gating
        if est_status not in ('draft', 'superseded') and not est_line_items:
            est_status = 'draft'

        self.add_fixture('estimates.estimate', estimate_pk, {
            'job': job_pk,
            'estimate_number': est_ref,
            'version': 1,
            'status': est_status,
            'parent': None,
            'created_date': created_date,
            'sent_date': created_date if est_status in ('open', 'accepted', 'rejected') else None,
            'closed_date': None,
            'expiration_date': None,
        })

        line_number = 1
        for item in est_line_items:
            li_pk = self.get_next_pk('estimates.estimatelineitem')
            qty = self._parse_decimal(item.get('Quantity', 1))
            price = self._parse_decimal(item.get('Price', 0))
            self.add_fixture('estimates.estimatelineitem', li_pk, {
                'estimate': estimate_pk,
                'task': None,
                'price_list_item': None,
                'line_number': line_number,
                'qty': str(qty),
                'units': '',
                'description': item.get('Description', '') or '',
                'price': str(price),
            })
            line_number += 1

        # Create WorkOrder + Tasks from estimate line items (if job is approved or completed)
        if job_status in ('approved', 'completed') and est_line_items:
            wo_pk = self.get_next_pk('jobs.workorder')
            wo_status = 'complete' if job_status == 'completed' else 'incomplete'

            self.add_fixture('jobs.workorder', wo_pk, {
                'job': job_pk,
                'status': wo_status,
                'template': None,
            })

            for task_num, item in enumerate(est_line_items, start=1):
                task_pk = self.get_next_pk('jobs.task')
                desc = item.get('Description', '') or ''
                task_name = desc[:255] if desc else f"Task {task_num}"
                task_status = 'complete' if wo_status == 'complete' else 'pending'

                self.add_fixture('jobs.task', task_pk, {
                    'parent_task': None,
                    'assignee': None,
                    'work_order': wo_pk,
                    'est_worksheet': None,
                    'name': task_name,
                    'sort_order': task_num,
                    'status': task_status,
                    'units': 'hours',
                    'rate': str(self._parse_decimal(item.get('Price', 0))),
                    'est_qty': str(self._parse_decimal(item.get('Quantity', 0))),
                })

        # Create Invoice(s) + line items
        for inv in invoices:
            inv_status = inv_status_map.get(inv.get('Status'), 'draft')
            paid_date = inv.get('Paid Date')
            if inv_status == 'open' and paid_date:
                paid_amount = float(inv.get('Paid Amount', 0) or 0)
                total_value = float(inv.get('Total Value', 0) or 0)
                if total_value and abs(paid_amount - total_value) < 0.01:
                    inv_status = 'paid'

            inv_line_items = inv.get('_line_items', [])
            if inv_status != 'draft' and not inv_line_items:
                inv_status = 'draft'

            inv_created = self._format_date(inv.get('Date'))
            inv_closed = self._format_date(paid_date) if inv_status == 'paid' else None

            inv_pk = self.get_next_pk('invoicing.invoice')
            self.add_fixture('invoicing.invoice', inv_pk, {
                'job': job_pk,
                'invoice_number': inv.get('Reference', '') or '',
                'status': inv_status,
                'created_date': inv_created,
                'sent_date': inv_created if inv_status in ('open', 'paid', 'partly-paid') else None,
                'closed_date': inv_closed,
            })

            ln = 1
            for item in inv_line_items:
                li_pk = self.get_next_pk('invoicing.invoicelineitem')
                qty = self._parse_decimal(item.get('Quantity', 1))
                price = self._parse_decimal(item.get('Price', 0))
                self.add_fixture('invoicing.invoicelineitem', li_pk, {
                    'invoice': inv_pk,
                    'task': None,
                    'price_list_item': None,
                    'line_number': ln,
                    'qty': str(qty),
                    'units': '',
                    'description': item.get('Description', '') or '',
                    'price': str(price),
                })
                ln += 1

        # Create PO + Bill for linked bills
        for bill in bills:
            bill_line_items = bill.get('_line_items', [])

            bill_date = self._format_date(bill.get('Date'))
            bill_year = bill.get('Date').year if isinstance(bill.get('Date'), datetime) else year

            if bill_year not in po_counters:
                po_counters[bill_year] = 0
            po_counters[bill_year] += 1
            po_number = f"PO{bill_year}-{po_counters[bill_year]:04d}"

            # Resolve bill contact/business
            bill_org = bill.get('Contact Organisation')
            bill_contact_name = bill.get('Contact Name')
            bill_contact_pk = None
            bill_biz_pk = None

            if bill_org:
                bill_biz_pk = self.business_map.get(bill_org)
                if bill_contact_name:
                    bill_contact_pk = self._resolve_contact(
                        org=bill_org, contact_name=bill_contact_name,
                        sheet='Bills', row=bill.get('_row', 0),
                        context=est_ref
                    )
            elif bill_contact_name:
                bill_biz_pk, bill_contact_pk = self._ensure_individual_vendor(bill_contact_name)

            if not bill_biz_pk:
                continue  # Skip bills we can't associate

            po_status = 'issued' if bill_line_items else 'draft'
            po_pk = self.get_next_pk('purchasing.purchaseorder')
            self.add_fixture('purchasing.purchaseorder', po_pk, {
                'po_number': po_number,
                'business': bill_biz_pk,
                'contact': bill_contact_pk,
                'status': po_status,
                'created_date': bill_date,
                'requested_date': None,
                'issued_date': bill_date if po_status == 'issued' else None,
                'received_date': None,
                'cancel_date': None,
            })

            if bill_year not in bill_counters:
                bill_counters[bill_year] = 0
            bill_counters[bill_year] += 1
            bill_number = f"B{bill_year}-{bill_counters[bill_year]:04d}"

            bill_pk = self.get_next_pk('purchasing.bill')
            self.add_fixture('purchasing.bill', bill_pk, {
                'bill_number': bill_number,
                'purchase_order': po_pk,
                'business': bill_biz_pk,
                'contact': bill_contact_pk,
                'vendor_invoice_number': bill.get('Reference', '') or '',
                'status': 'draft',
                'created_date': bill_date,
                'due_date': self._format_date(bill.get('Due Date')),
                'received_date': None,
                'paid_date': None,
                'cancelled_date': None,
            })

            ln = 1
            for item in bill_line_items:
                qty = self._parse_decimal(item.get('Quantity', 1)) or Decimal('1')
                net_value = abs(self._parse_decimal(item.get('Net Value', 0)))
                price = net_value / qty

                po_li_pk = self.get_next_pk('purchasing.purchaseorderlineitem')
                self.add_fixture('purchasing.purchaseorderlineitem', po_li_pk, {
                    'purchase_order': po_pk,
                    'job': job_pk,
                    'task': None, 'price_list_item': None,
                    'line_number': ln,
                    'qty': str(qty),
                    'units': item.get('Item Type', '-no unit-') or '-no unit-',
                    'description': item.get('Description', '') or '',
                    'price': str(price),
                })

                bill_li_pk = self.get_next_pk('purchasing.billlineitem')
                self.add_fixture('purchasing.billlineitem', bill_li_pk, {
                    'bill': bill_pk,
                    'task': None, 'price_list_item': None,
                    'line_number': ln,
                    'qty': str(qty),
                    'units': item.get('Item Type', '-no unit-') or '-no unit-',
                    'description': item.get('Description', '') or '',
                    'price': str(price),
                })
                ln += 1

        return job_pk

    def _build_implicit_jobs(self):
        """Create Jobs for estimates linked to invoices but not to any project."""
        import re

        if self.verbose:
            print("  Building implicit jobs from unlinked estimates...")

        # Parse all records
        all_estimates = self._collect_all_estimates()
        all_invoices = self._collect_all_invoices()
        all_bills = self._collect_all_bills()

        # Build set of already-processed estimate refs
        processed_est_refs = set()
        for f in self.fixture_data:
            if f['model'] == 'estimates.estimate':
                processed_est_refs.add(f['fields'].get('estimate_number'))

        # Build set of already-processed invoice refs
        processed_inv_refs = set()
        for f in self.fixture_data:
            if f['model'] == 'invoicing.invoice':
                processed_inv_refs.add(f['fields'].get('invoice_number'))

        # Index unlinked invoices by reference
        invoice_by_ref = {}
        for inv in all_invoices:
            ref = inv.get('Reference')
            if ref and ref not in processed_inv_refs:
                invoice_by_ref[ref] = inv

        # Index unlinked bills by estimate ref found in Comments
        bills_by_est_ref = {}
        for bill in all_bills:
            comment = bill.get('Comments', '') or ''
            matches = re.findall(r'\b(\d{5})\b', comment)
            for m in matches:
                bills_by_est_ref.setdefault(m, []).append(bill)

        # Build deposit invoice index: estimate_ref → [invoices]
        deposit_invoices_by_est = {}
        for inv in all_invoices:
            ref = inv.get('Reference')
            if ref in processed_inv_refs:
                continue
            for item in inv.get('_line_items', []):
                desc = (item.get('Description', '') or '').lower()
                matches = re.findall(r'\b(\d{5})\b', desc)
                for m in matches:
                    deposit_invoices_by_est.setdefault(m, []).append(inv)
                    break  # One match per invoice is enough

        # Initialize counters from existing fixture data
        job_counters = {}
        po_counters = {}
        bill_counters = {}
        for f in self.fixture_data:
            if f['model'] == 'jobs.job':
                jn = f['fields'].get('job_number', '')
                if jn.startswith('J'):
                    try:
                        parts = jn.split('-')
                        y, c = int(parts[0][1:]), int(parts[1])
                        job_counters[y] = max(job_counters.get(y, 0), c)
                    except (ValueError, IndexError):
                        pass
            elif f['model'] == 'purchasing.purchaseorder':
                pn = f['fields'].get('po_number', '')
                if pn.startswith('PO'):
                    try:
                        parts = pn.split('-')
                        y, c = int(parts[0][2:]), int(parts[1])
                        po_counters[y] = max(po_counters.get(y, 0), c)
                    except (ValueError, IndexError):
                        pass
            elif f['model'] == 'purchasing.bill':
                bn = f['fields'].get('bill_number', '')
                if bn.startswith('B') and '-' in bn:
                    try:
                        parts = bn.split('-')
                        y, c = int(parts[0][1:]), int(parts[1])
                        bill_counters[y] = max(bill_counters.get(y, 0), c)
                    except (ValueError, IndexError):
                        pass

        # Process unlinked estimates that have an Invoice Reference
        implicit_count = 0
        used_invoice_refs = set()

        for est in all_estimates:
            est_ref = est.get('Reference')
            if not est_ref or est_ref in processed_est_refs:
                continue

            # Find linked invoices
            inv_ref = est.get('Invoice Reference')
            linked_invoices = []
            if inv_ref and inv_ref in invoice_by_ref and inv_ref not in used_invoice_refs:
                linked_invoices.append(invoice_by_ref[inv_ref])
                used_invoice_refs.add(inv_ref)

            # Also check deposit invoices
            for dep_inv in deposit_invoices_by_est.get(est_ref, []):
                dep_ref = dep_inv.get('Reference')
                if dep_ref and dep_ref not in used_invoice_refs:
                    linked_invoices.append(dep_inv)
                    used_invoice_refs.add(dep_ref)

            if not linked_invoices:
                continue  # No invoices — may be handled by _build_recent_unlinked_estimates

            # Find linked bills
            linked_bills = bills_by_est_ref.get(est_ref, [])

            job_pk = self._create_implicit_job(
                est, linked_invoices, linked_bills,
                job_counters, po_counters, bill_counters,
            )
            if job_pk:
                implicit_count += 1
                processed_est_refs.add(est_ref)

        # Store processed refs for _build_recent_unlinked_estimates
        self._processed_est_refs = processed_est_refs

        if self.verbose:
            print(f"    Created {implicit_count} implicit jobs")

    def _build_recent_unlinked_estimates(self):
        """Create draft Jobs for recent unlinked estimates (< 6 months old, not rejected)."""
        if self.verbose:
            print("  Building jobs for recent unlinked estimates...")

        cutoff_date = datetime(2025, 10, 1)

        all_estimates = self._collect_all_estimates()
        processed = getattr(self, '_processed_est_refs', set())

        est_status_map = {
            'Draft': 'draft', 'Sent': 'open',
            'Approved': 'accepted', 'Rejected': 'rejected',
        }

        # Initialize job counters from existing data
        job_counters = {}
        for f in self.fixture_data:
            if f['model'] == 'jobs.job':
                jn = f['fields'].get('job_number', '')
                if jn.startswith('J'):
                    try:
                        parts = jn.split('-')
                        y, c = int(parts[0][1:]), int(parts[1])
                        job_counters[y] = max(job_counters.get(y, 0), c)
                    except (ValueError, IndexError):
                        pass

        count = 0
        for est in all_estimates:
            est_ref = est.get('Reference')
            if not est_ref or est_ref in processed:
                continue

            est_status = est_status_map.get(est.get('Status'), 'draft')
            if est_status == 'rejected':
                continue

            est_date = est.get('Date')
            if not est_date or not isinstance(est_date, datetime) or est_date < cutoff_date:
                continue

            est_line_items = est.get('_line_items', [])

            # Resolve contact
            contact_pk, business_pk = self._resolve_or_create_contact_for_estimate(est)
            if not contact_pk:
                continue

            # Job status from estimate status
            if est_status == 'accepted':
                job_status = 'approved'
            elif est_status == 'open':
                job_status = 'submitted'
            else:
                job_status = 'draft'

            created_date = self._format_date(est_date)
            year = est_date.year

            if year not in job_counters:
                job_counters[year] = 0
            job_counters[year] += 1
            job_number = f"J{year}-{job_counters[year]:04d}"

            # Job name
            first_desc = ''
            if est_line_items:
                first_desc = (est_line_items[0].get('Description', '') or '')[:40]
            job_name = f"Est {est_ref}"
            if first_desc:
                job_name = f"Est {est_ref} - {first_desc}"
            job_name = job_name[:50]

            start_date = created_date if job_status == 'approved' else None

            job_pk = self.get_next_pk('jobs.job')
            self.add_fixture('jobs.job', job_pk, {
                'name': job_name,
                'job_number': job_number,
                'contact': contact_pk,
                'start_date': start_date,
                'due_date': None,
                'created_date': created_date,
                'customer_po_number': '',
                'status': job_status,
                'description': est.get('Notes', '') or '',
                'completed_date': None,
            })

            # Create Estimate + line items
            estimate_pk = self.get_next_pk('estimates.estimate')

            if est_status not in ('draft', 'superseded') and not est_line_items:
                est_status = 'draft'

            self.add_fixture('estimates.estimate', estimate_pk, {
                'job': job_pk,
                'estimate_number': est_ref,
                'version': 1,
                'status': est_status,
                'parent': None,
                'created_date': created_date,
                'sent_date': created_date if est_status in ('open', 'accepted') else None,
                'closed_date': None,
                'expiration_date': None,
            })

            line_number = 1
            for item in est_line_items:
                li_pk = self.get_next_pk('estimates.estimatelineitem')
                qty = self._parse_decimal(item.get('Quantity', 1))
                price = self._parse_decimal(item.get('Price', 0))
                self.add_fixture('estimates.estimatelineitem', li_pk, {
                    'estimate': estimate_pk,
                    'task': None,
                    'price_list_item': None,
                    'line_number': line_number,
                    'qty': str(qty),
                    'units': '',
                    'description': item.get('Description', '') or '',
                    'price': str(price),
                })
                line_number += 1

            # If approved, create WorkOrder + Tasks from line items
            if job_status == 'approved' and est_line_items:
                wo_pk = self.get_next_pk('jobs.workorder')
                self.add_fixture('jobs.workorder', wo_pk, {
                    'job': job_pk,
                    'status': 'incomplete',
                    'template': None,
                })

                for task_num, item in enumerate(est_line_items, start=1):
                    task_pk = self.get_next_pk('jobs.task')
                    desc = item.get('Description', '') or ''
                    task_name = desc[:255] if desc else f"Task {task_num}"

                    self.add_fixture('jobs.task', task_pk, {
                        'parent_task': None,
                        'assignee': None,
                        'work_order': wo_pk,
                        'est_worksheet': None,
                        'name': task_name,
                        'sort_order': task_num,
                        'status': 'pending',
                        'units': 'hours',
                        'rate': str(self._parse_decimal(item.get('Price', 0))),
                        'est_qty': str(self._parse_decimal(item.get('Quantity', 0))),
                    })

            count += 1
            processed.add(est_ref)

        if self.verbose:
            print(f"    Created {count} jobs from recent unlinked estimates")

    def _reconcile_states(self):
        """
        Reconcile cross-model state constraints after all objects are built.

        Enforces:
        - Estimate sent → Job at least submitted (constraint #2)
        - Estimate accepted → Job at least approved (constraint #2 + #6)
        - All invoices paid → Job completed (constraint #3)
        - Sent invoice similar to estimate → tasks complete
        """
        if self.verbose:
            print("  Reconciling cross-model states...")

        # Index fixtures by model for lookup
        estimates_by_job = {}  # job_pk -> [fixture]
        invoices_by_job = {}   # job_pk -> [fixture]
        job_fixtures = {}      # job_pk -> fixture
        est_line_items = {}    # estimate_pk -> [fixture]
        inv_line_items = {}    # invoice_pk -> [fixture]
        wo_by_job = {}         # job_pk -> wo_pk
        tasks_by_wo = {}       # wo_pk -> [fixture]

        for fixture in self.fixture_data:
            model = fixture['model']
            if model == 'estimates.estimate':
                job_pk = fixture['fields']['job']
                estimates_by_job.setdefault(job_pk, []).append(fixture)
            elif model == 'invoicing.invoice':
                job_pk = fixture['fields']['job']
                invoices_by_job.setdefault(job_pk, []).append(fixture)
            elif model == 'jobs.job':
                job_fixtures[fixture['pk']] = fixture
            elif model == 'estimates.estimatelineitem':
                est_pk = fixture['fields']['estimate']
                est_line_items.setdefault(est_pk, []).append(fixture)
            elif model == 'invoicing.invoicelineitem':
                inv_pk = fixture['fields']['invoice']
                inv_line_items.setdefault(inv_pk, []).append(fixture)
            elif model == 'jobs.workorder':
                job_pk = fixture['fields']['job']
                wo_by_job[job_pk] = fixture['pk']
            elif model == 'jobs.task':
                wo_pk = fixture['fields']['work_order']
                if wo_pk:
                    tasks_by_wo.setdefault(wo_pk, []).append(fixture)

        changes = 0

        for job_pk, job_fix in job_fixtures.items():
            fields = job_fix['fields']
            status = fields['status']

            # Skip terminal statuses
            if status in ('completed', 'cancelled', 'rejected'):
                continue

            estimates = estimates_by_job.get(job_pk, [])
            invoices = invoices_by_job.get(job_pk, [])

            # Constraint #2: If any estimate has been sent (open or later),
            # the job must be at least submitted
            has_sent = any(
                e['fields']['status'] in ('open', 'accepted', 'rejected',
                                          'expired', 'superseded')
                for e in estimates
            )
            if has_sent and status == 'draft':
                fields['status'] = 'submitted'
                status = 'submitted'
                changes += 1
                if self.verbose:
                    print(f"    Job {fields.get('job_number')}: draft → submitted (has sent estimate)")

            # Constraint #2 continued: If any estimate is accepted,
            # the job must be at least approved
            has_accepted = any(
                e['fields']['status'] == 'accepted' for e in estimates
            )
            if has_accepted and status in ('draft', 'submitted'):
                fields['status'] = 'approved'
                if not fields.get('start_date'):
                    # Use the accepted estimate's sent_date or created_date
                    for est in estimates:
                        if est['fields']['status'] == 'accepted':
                            fields['start_date'] = (
                                est['fields'].get('sent_date')
                                or est['fields'].get('created_date')
                            )
                            break
                status = 'approved'
                changes += 1
                if self.verbose:
                    print(f"    Job {fields.get('job_number')}: → approved (has accepted estimate)")

            # Constraint #3: If all invoices for the job are paid,
            # the job must be completed
            if invoices and all(
                inv['fields']['status'] == 'paid' for inv in invoices
            ):
                if status in ('draft', 'submitted', 'approved'):
                    fields['status'] = 'completed'
                    # Set completed_date from the last invoice's closed_date
                    closed_dates = [
                        inv['fields']['closed_date']
                        for inv in invoices
                        if inv['fields'].get('closed_date')
                    ]
                    if closed_dates:
                        fields['completed_date'] = max(closed_dates)
                    status = 'completed'
                    changes += 1
                    if self.verbose:
                        print(f"    Job {fields.get('job_number')}: → completed (all invoices paid)")

        # Expire old open estimates: if created_date is more than 30 days ago,
        # an estimate cannot remain 'open' — move it to 'expired'.
        # When the latest estimate on a job expires, reject the job.
        expire_days = 30
        today = date.today()
        expired_estimates = 0
        rejected_jobs = 0

        for job_pk, estimates in estimates_by_job.items():
            for est in estimates:
                est_fields = est['fields']
                if est_fields['status'] != 'open':
                    continue
                created = est_fields.get('created_date')
                if not created:
                    continue
                created_date = date.fromisoformat(created) if isinstance(created, str) else created
                if (today - created_date).days > expire_days:
                    est_fields['status'] = 'expired'
                    est_fields['closed_date'] = today.isoformat()
                    expired_estimates += 1
                    if self.verbose:
                        print(f"    Estimate {est['pk']}: open → expired (older than {expire_days} days)")

            # After expiring estimates, check if the job should be rejected.
            # If all non-superseded estimates are now in terminal states
            # (expired/rejected) and none are accepted, reject the job.
            job_fix = job_fixtures.get(job_pk)
            if not job_fix:
                continue
            job_status = job_fix['fields']['status']
            if job_status in ('completed', 'cancelled', 'rejected'):
                continue
            non_superseded = [e for e in estimates if e['fields']['status'] != 'superseded']
            if non_superseded and all(
                e['fields']['status'] in ('expired', 'rejected') for e in non_superseded
            ):
                job_fix['fields']['status'] = 'rejected'
                rejected_jobs += 1
                if self.verbose:
                    print(f"    Job {job_fix['fields'].get('job_number')}: → rejected (all estimates expired/rejected)")

        if self.verbose:
            print(f"    Expired {expired_estimates} open estimates (>{expire_days} days old)")
            print(f"    Rejected {rejected_jobs} jobs (all estimates expired/rejected)")

        # Sent invoice matches estimate → complete all tasks on the job.
        # "Substantially similar" = total values within 10%.
        def _line_item_total(items):
            total = Decimal('0')
            for li in items:
                f = li['fields']
                total += Decimal(f.get('qty', '0')) * Decimal(f.get('price', '0'))
            return total

        tasks_completed = 0
        jobs_with_completed_tasks = 0
        for job_pk in job_fixtures:
            estimates = estimates_by_job.get(job_pk, [])
            invoices = invoices_by_job.get(job_pk, [])
            if not estimates or not invoices:
                continue

            # Check for any sent invoice (non-draft)
            sent_invoices = [
                inv for inv in invoices if inv['fields']['status'] != 'draft'
            ]
            if not sent_invoices:
                continue

            # Get estimate totals (use latest non-superseded estimate)
            active_estimates = [
                e for e in estimates if e['fields']['status'] != 'superseded'
            ]
            if not active_estimates:
                active_estimates = estimates

            est_totals = []
            for est in active_estimates:
                items = est_line_items.get(est['pk'], [])
                if items:
                    est_totals.append(_line_item_total(items))

            if not est_totals:
                continue

            # Check if any sent invoice total is within 10% of any estimate total
            match_found = False
            for inv in sent_invoices:
                inv_items = inv_line_items.get(inv['pk'], [])
                if not inv_items:
                    continue
                inv_total = _line_item_total(inv_items)
                if inv_total == 0:
                    continue
                for est_total in est_totals:
                    if est_total == 0:
                        continue
                    ratio = inv_total / est_total
                    if Decimal('0.9') <= ratio <= Decimal('1.1'):
                        match_found = True
                        break
                if match_found:
                    break

            if not match_found:
                continue

            # Complete all tasks on this job's work order
            wo_pk = wo_by_job.get(job_pk)
            if not wo_pk:
                continue
            tasks = tasks_by_wo.get(wo_pk, [])
            if not tasks:
                continue

            job_touched = False
            for task_fix in tasks:
                if task_fix['fields']['status'] != 'complete':
                    task_fix['fields']['status'] = 'complete'
                    tasks_completed += 1
                    job_touched = True
            if job_touched:
                jobs_with_completed_tasks += 1
                if self.verbose:
                    job_num = job_fixtures[job_pk]['fields'].get('job_number')
                    print(f"    Job {job_num}: completed tasks (sent invoice matches estimate)")

        if self.verbose:
            print(f"    Completed {tasks_completed} tasks across {jobs_with_completed_tasks} jobs (invoice ≈ estimate)")

        # Cancel old approved jobs: if the job started before 2026 and is
        # still in 'approved' status, it's stale — mark it cancelled.
        cancelled = 0
        for job_pk, job_fix in job_fixtures.items():
            fields = job_fix['fields']
            if fields['status'] != 'approved':
                continue
            # Use start_date if available, otherwise created_date
            ref_date = fields.get('start_date') or fields.get('created_date')
            if ref_date and ref_date < '2026-01-01':
                fields['status'] = 'cancelled'
                cancelled += 1
                if self.verbose:
                    print(f"    Job {fields.get('job_number')}: approved → cancelled (started before 2026)")

        if self.verbose:
            print(f"    Reconciled {changes} job statuses")
            print(f"    Cancelled {cancelled} old approved jobs (pre-2026)")

    def _build_configuration(self):
        """
        Build Configuration fixtures for all app settings.

        Scans generated fixtures to compute document number counters, then
        emits sequence patterns matching the formats used in this script,
        plus all other configuration keys the app expects.
        """
        if self.verbose:
            print("  Building configuration...")

        # Remove any Configuration entries that came from base fixtures —
        # we'll replace them with computed values.
        self.fixture_data = [
            f for f in self.fixture_data
            if f['model'] != 'core.configuration'
        ]

        # --- Compute max counters from generated document numbers ---
        # The script uses per-year counters but the app uses a single global
        # counter, so we take the max counter value across all years.
        counter_patterns = {
            'job': ('jobs.job', 'job_number', r'J(\d{4})-(\d+)'),
            'po': ('purchasing.purchaseorder', 'po_number', r'PO(\d{4})-(\d+)'),
            'bill': ('purchasing.bill', 'bill_number', r'B(\d{4})-(\d+)'),
        }

        max_counters = {doc_type: 0 for doc_type in counter_patterns}

        for fixture in self.fixture_data:
            for doc_type, (model, field, pattern) in counter_patterns.items():
                if fixture['model'] == model:
                    value = fixture['fields'].get(field, '')
                    match = re.match(pattern, value)
                    if match:
                        counter = int(match.group(2))
                        max_counters[doc_type] = max(max_counters[doc_type], counter)

        # Estimates and invoices use spreadsheet reference numbers, not
        # generated sequences — start their counters at 0.
        max_counters['estimate'] = 0
        max_counters['invoice'] = 0

        # --- Document numbering sequences ---
        # Patterns match the formats this script generates
        sequences = {
            'job_number_sequence': 'J{year}-{counter:04d}',
            'job_counter': str(max_counters['job']),
            'estimate_number_sequence': 'EST-{year}-{counter:04d}',
            'estimate_counter': str(max_counters['estimate']),
            'invoice_number_sequence': 'INV-{year}-{counter:04d}',
            'invoice_counter': str(max_counters['invoice']),
            'po_number_sequence': 'PO{year}-{counter:04d}',
            'po_counter': str(max_counters['po']),
            'bill_number_sequence': 'B{year}-{counter:04d}',
            'bill_counter': str(max_counters['bill']),
        }

        # --- Other configuration keys ---
        other_configs = {
            'default_tax_rate': '0.0825',
            'est_expire_days': '30',
            'board_closed_retention_days': '14',
            'email_retention_days': '90',
            'email_display_limit': '30',
            'units_list': json.dumps([
                "none", "hours", "ea", "sq ft", "ft", "yd", "m",
                "sheets", "pcs", "lbs", "kg", "gal", "qt", "L",
                "bd ft", "ln ft"
            ]),
        }

        all_configs = {**sequences, **other_configs}

        for key, value in all_configs.items():
            self.add_fixture('core.configuration', key, {'value': value})

        if self.verbose:
            print(f"    Created {len(all_configs)} configuration entries")
            for doc_type in ('job', 'estimate', 'invoice', 'po', 'bill'):
                print(f"    {doc_type} counter: {max_counters[doc_type]}")

    @staticmethod
    def _split_name(full_name: str) -> Tuple[str, str]:
        """Split 'First Last' into (first_name, last_name)."""
        parts = full_name.strip().split(None, 1)
        if len(parts) == 2:
            return parts[0], parts[1]
        elif len(parts) == 1:
            return parts[0], '(unknown)'
        return '(unknown)', '(unknown)'

    def _format_date(self, value) -> Optional[str]:
        """Format date value to ISO string."""
        if not value:
            return None

        if isinstance(value, datetime):
            return value.date().isoformat()

        return None

    def _parse_decimal(self, value) -> Decimal:
        """Parse decimal value."""
        if value is None:
            return Decimal('0')

        try:
            return Decimal(str(value))
        except:
            return Decimal('0')

    def _parse_revision_suffix(self, reference: str) -> tuple:
        """
        Parse revision suffix from reference number.

        Examples:
            'EST123' -> ('EST123', 1)
            'EST123-r2' -> ('EST123', 2)
            'EST123-r3' -> ('EST123', 3)
            'EST123-rev2' -> ('EST123', 2)

        Returns:
            (base_reference, revision_number)
        """
        import re

        if not reference:
            return ('', 1)

        # Match patterns like -r2, -r3, -rev2, -rev3
        match = re.search(r'-r(?:ev)?(\d+)$', reference, re.IGNORECASE)

        if match:
            revision = int(match.group(1))
            base_ref = reference[:match.start()]
            return (base_ref, revision)

        # No revision suffix found
        return (reference, 1)

    def _print_summary(self):
        """Print summary of objects to be created."""
        print("=" * 70)
        print("Summary")
        print("=" * 70)

        model_counts = {}
        for fixture in self.fixture_data:
            model = fixture['model']
            model_counts[model] = model_counts.get(model, 0) + 1

        print("\nObjects to be created:")
        for model in sorted(model_counts.keys()):
            print(f"  {model:40} {model_counts[model]:5}")

        print(f"\n  {'TOTAL':40} {len(self.fixture_data):5}")

    def _write_json(self):
        """Write fixture data to JSON file."""
        with open(self.output_path, 'w') as f:
            json.dump(self.fixture_data, f, indent=2, default=str)


def main():
    parser = argparse.ArgumentParser(
        description='Convert Neal\'s CNC Excel export to Django fixture JSON',
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Examples:
  %(prog)s nealsdata/company-export.xlsx
  %(prog)s nealsdata/company-export.xlsx --output my_data.json
  %(prog)s nealsdata/company-export.xlsx --non-interactive
  %(prog)s nealsdata/company-export.xlsx --dry-run --verbose
        """
    )

    parser.add_argument(
        'excel_file',
        help='Path to Excel file to convert'
    )

    parser.add_argument(
        '--output',
        default='neals_data.json',
        help='Output JSON file path (default: neals_data.json)'
    )

    parser.add_argument(
        '--non-interactive',
        action='store_true',
        help='Skip interactive prompts, auto-map all contact mismatches'
    )

    parser.add_argument(
        '--dry-run',
        action='store_true',
        help='Show statistics without generating JSON file'
    )

    parser.add_argument(
        '--verbose',
        action='store_true',
        help='Show detailed progress information'
    )

    args = parser.parse_args()

    # Check if file exists
    if not Path(args.excel_file).exists():
        print(f"Error: File not found: {args.excel_file}")
        sys.exit(1)

    # Run converter
    converter = NealsDataConverter(
        excel_path=args.excel_file,
        output_path=args.output,
        interactive=not args.non_interactive,
        dry_run=args.dry_run,
        verbose=args.verbose
    )

    try:
        converter.convert()
    except KeyboardInterrupt:
        print("\n\nConversion cancelled by user.")
        sys.exit(1)
    except Exception as e:
        print(f"\n\nError during conversion: {e}")
        if args.verbose:
            import traceback
            traceback.print_exc()
        sys.exit(1)


if __name__ == '__main__':
    main()
